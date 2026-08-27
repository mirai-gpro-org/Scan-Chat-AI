#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LAiF 上り入力フォーム（AI疾病発症予測）生成 — 決定論ビルダー。

役割: 健診(HealthCheckupData) ＋ AI問診(LifestyleQuestionnaireData) ＋ 基本情報 を、
      LAiF提供の正式フォーム `input_format_new_202312.xlsx`(No.0〜157) の「入力フィールド」列(G)へ写像する。
正本仕様: docs/subscription/kit_lifecycle_and_handoff_management_spec.md §4.1.1
          docs/lab/laif_s3_secure_handoff_spec.md

確定ルール(LAiF回答 2026-08):
  ・**名前欄(G1)＝識別番号(仮名ID)を必ずセット**(空欄不可＝JOBRUNNER解析不能)。実氏名は載せない=PII非送付。
  ・No.0 識別番号 も同一の仮名ID。
  ・**識別番号は英字を1文字以上含む**(数字のみは解析不可・大小不問。LAiF回答 2026-08-26)。

記入ルールはテンプレート自身から読む (2026-08-27・LAiF指摘を受けた改修):
  フォームには入力規則(プルダウン/数値範囲)が24個定義済みで、これが唯一の正解。
  備考欄(H列)は表記がゆれている(プルダウン `(-)` に対し備考 `(ー)`)ため手本にしない。
  → `scripts/laif_form_rules.py` が 規則読取/値スナップ/単位換算/親子ルール/全数検証 を担う。
  出力前に全セルを検証し、**違反が残れば書き出さずに異常終了する**。
捏造ゼロ:
  ・与えられた実値のみ記入。写像先が無い/値が無い項目は空のまま(推定で埋めない)。
  ・写像は項目名の正規化一致(＋別名辞書)で決定論。未一致は unmatched として報告(黙って捨てない)。

使い方:
  python3 scripts/build-laif-input.py --template <正式フォーム.xlsx> --input <data.json> --out <出力.xlsx>
  data.json:
    {
      "client_id": "W026000123",                      # 識別番号(仮名ID・No.0/名前欄に入る・**英字必須**)
      "profile": {"exam_date":"20260805","sex":"男性","birth_date":"19720101",
                  "age":"54","height":"171.2","weight":"68.4","bmi":"23.3","waist":"84.0"},
      "measurements": [{"name":"白血球数","value":"5800","unit":"/μL"}, {"name":"リンパ球","value":"34.0"}, ...],
        # 白血球数は /μL・x10^2/μL・x10^3/μL のどれで来ても x10^3/μL へ自動換算される
      "questionnaire": {"朝食抜き":"N", "睡眠で休養が取れている":"Y", ...}   # LAiF項目名(正規化一致)
    }
"""
import argparse, json, re, sys, os, unicodedata
from collections import Counter
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from laif_form_rules import (  # noqa: E402
    read_rules, normalize_value, validate, parent_child_map,
    check_client_id, item_name, COL_INPUT,
)

def norm_exact(s):
    """厳密一致用: NFKC(全角括弧→半角等)+空白除去+小文字化。**括弧の中身は残す**。"""
    s = unicodedata.normalize('NFKC', str(s if s is not None else ''))
    return re.sub(r'[\s\u3000]', '', s).lower()


def norm(s):
    """緩い一致用: 括弧内(英略称等)を除去。**同名多発(既往歴(…)等)があるため単独では使わない**。"""
    if s is None:
        return ''
    s = re.sub(r'[（(].*?[）)]', '', str(s))
    return re.sub(r'[\s\u3000]', '', s).lower()

# 別名(正規化後キー → LAiF正規化名)。非自明な同義のみ。
ALIAS = {
    'ast': 'got', 'gpt': 'got',  # measurement 'AST'→LAiF 'GOT（AST）'(norm 'got') / 'ALT'→'GPT'
    'alt': 'gpt',
    'γ-gtp': 'γ-gtp', 'γgtp': 'γ-gtp', 'gammagtp': 'γ-gtp',
    'egfr': 'egrf', 'egrf': 'egrf',      # LAiF綴りは 'eGRF'
    '中性脂肪': '中性脂肪', 'トリグリセライド': '中性脂肪',
    'クレアチニン': 'クレアチニン',
    '尿素窒素': '尿酸窒素',              # LAiF表記ゆれ 'BUN'=尿酸窒素(原文ママ)
    'bun': '尿酸窒素',
    '空腹時血糖': '空腹時血糖値', '血糖': '空腹時血糖値',
    'alp': 'アルカリフォスターゼ', 'アルカリホスファターゼ': 'アルカリフォスターゼ', 'アルカリフォスファターゼ': 'アルカリフォスターゼ',
    # 血圧: スキャン出力/一般名 → LAiF「最高血圧 初回値/最低血圧 初回値」(norm後は空白除去済)
    '収縮期血圧': '最高血圧初回値', '最高血圧': '最高血圧初回値', '血圧最高': '最高血圧初回値', 'sbp': '最高血圧初回値',
    '拡張期血圧': '最低血圧初回値', '最低血圧': '最低血圧初回値', '血圧最低': '最低血圧初回値', 'dbp': '最低血圧初回値',
}

BASIC = {  # profile キー → LAiF項目(正規化名)
    'exam_date': '受診日', 'sex': '性別', 'birth_date': '生年月日', 'age': '年齢',
    'height': '身長', 'weight': '体重', 'bmi': 'bmi', 'waist': '腹囲',
}

def _report(ws, notes, violations, audit, out_path, warns):
    print('=== build-laif-input 監査 ===', file=sys.stderr)
    if audit:
        print(f'名前欄/識別番号 : {audit["name_id"]}', file=sys.stderr)
        print(f'基本情報 記入   : {audit["basic"]}', file=sys.stderr)
        print(f'健診 一致       : {len(audit["measurements_matched"])}  未一致: {len(audit["measurements_unmatched"])}',
              file=sys.stderr)
        if audit['measurements_unmatched']:
            print('  未一致(健診): ' + ', '.join(map(str, audit['measurements_unmatched'])), file=sys.stderr)
        print(f'問診 一致       : {audit["questionnaire_matched"]}  未一致: {len(audit["questionnaire_unmatched"])}',
              file=sys.stderr)
        if audit['questionnaire_unmatched']:
            print('  未一致(問診): ' + ', '.join(map(str, audit['questionnaire_unmatched'])), file=sys.stderr)
    for w in warns:
        print(f'  警告: {w}', file=sys.stderr)
    if notes:
        print(f'正規化 {len(notes)} 件:', file=sys.stderr)
        for r, nm, note in notes:
            print(f'  行{r:>3} {nm[:22]:<24} {note}', file=sys.stderr)
    if violations:
        print(f'*** 入力規則違反 {len(violations)} 件 — 出力しません ***', file=sys.stderr)
        for v in violations:
            print(f'  行{v["row"]:>3} {v["name"][:22]:<24} [{v["kind"]}] {v["value"]!r} — {v["message"]}',
                  file=sys.stderr)
        return False
    print(f'入力規則: 違反なし ✅', file=sys.stderr)
    if out_path:
        print(f'出力: {out_path}', file=sys.stderr)
    return True


def _apply_parent_child(ws, rules, notes):
    """親（有/無）が「無」なら、その配下の子行（「有」のみ）を空にする。テンプレート仕様。"""
    for parent, children in parent_child_map(ws, rules).items():
        pv = str(ws.cell(parent, COL_INPUT).value or '').strip()
        if pv != '無':
            continue
        cleared = 0
        for c in children:
            if str(ws.cell(c, COL_INPUT).value or '').strip() != '':
                ws.cell(c, COL_INPUT).value = None
                cleared += 1
        if cleared:
            notes.append((parent, item_name(ws, parent),
                          f'親が「無」のため子 {cleared} 行を空にした（子は「有」しか選べない仕様）'))


def _open(path):
    wb = openpyxl.load_workbook(path)
    ws = wb['KM'] if 'KM' in wb.sheetnames else wb.worksheets[0]
    return wb, ws


def cmd_check(path):
    """既存ファイルを入力規則で全数検証する（書き換えない）。"""
    wb, ws = _open(path)
    cid = ws.cell(1, COL_INPUT).value
    errs, warns = check_client_id(cid)
    violations = validate(ws)
    print(f'=== check: {path} ===', file=sys.stderr)
    print(f'名前欄(G1) : {cid!r}', file=sys.stderr)
    for e in errs:
        print(f'  ERROR: {e}', file=sys.stderr)
    ok = _report(ws, [], violations, None, None, warns)
    return 0 if (ok and not errs) else 1


def cmd_repair(src, out, client_id=None):
    """既存ファイルを、ビルダーと同一の正規化ロジックで修正して書き出す。"""
    wb, ws = _open(src)
    rules = read_rules(ws)
    if client_id:
        ws.cell(1, COL_INPUT).value = client_id
        for r in range(3, ws.max_row + 1):
            if norm_exact(ws.cell(r, 4).value or ws.cell(r, 3).value or '') == '識別番号':
                ws.cell(r, COL_INPUT).value = client_id
                break
    notes = []
    for r in sorted(rules):
        v = ws.cell(r, COL_INPUT).value
        if v is None or str(v).strip() == '':
            continue
        nv, note = normalize_value(ws, r, v, rules)
        if nv != v:
            ws.cell(r, COL_INPUT).value = nv
        if note:
            notes.append((r, item_name(ws, r), note))
    _apply_parent_child(ws, rules, notes)
    cid = ws.cell(1, COL_INPUT).value
    errs, warns = check_client_id(cid)
    for e in errs:
        print(f'  ERROR(識別番号): {e}', file=sys.stderr)
    violations = validate(ws, rules)
    if not _report(ws, notes, violations, None, out, warns) or errs:
        return 1
    wb.save(out)
    print(f'修復出力: {out}', file=sys.stderr)
    return 0


def cmd_build(a):
    data = json.load(open(a.input, encoding='utf-8'))
    client_id = str(data.get('client_id') or '').strip()
    errs, warns = check_client_id(client_id)
    if errs:
        for e in errs:
            print(f'ERROR: {e}', file=sys.stderr)
        return 2

    wb, ws = _open(a.template)
    rules = read_rules(ws)
    notes = []

    # フィールド行の索引を 2 系統で作る。No列は数式のため項目名(C/D列)で検出。
    #   exact_row … 括弧の中身まで含めた厳密一致（既往歴（糖尿病）と既往歴（高血圧）を区別する）
    #   loose_row … 括弧を落とした一致。**曖昧でない(1行しか該当しない)ものだけ**採用。
    # 旧実装は loose のみで、既往歴/服薬情報/眼圧左右/便潜血1・2日目など
    # **80 行が先頭 1 行へ吸われて黙って取り違えられていた**（2026-08-27 修正）。
    SKIP = {'項目', 'no'}
    labels = []
    for r in range(3, ws.max_row + 1):
        c2 = ws.cell(r, 2).value
        if isinstance(c2, str) and c2.startswith('＜'):
            continue
        name = (ws.cell(r, 4).value or ws.cell(r, 3).value or '')
        if not norm_exact(name) or norm_exact(name) in SKIP:
            continue
        labels.append((r, name))
    exact_row = {}
    for r, name in labels:
        exact_row.setdefault(norm_exact(name), r)
    cnt = Counter(norm(name) for _, name in labels)
    loose_row = {}
    for r, name in labels:
        n = norm(name)
        if cnt[n] == 1:
            loose_row.setdefault(n, r)

    def resolve_row(raw_name):
        e, l = norm_exact(raw_name), norm(raw_name)
        if e in exact_row:
            return exact_row[e]
        a = ALIAS.get(l) or ALIAS.get(e)
        if a and (a in loose_row or a in exact_row):
            return loose_row.get(a) or exact_row.get(a)
        return loose_row.get(l)

    def put(raw_name, value):
        """値をテンプレートの入力規則へ正規化してから書き込む（素通ししない）。"""
        if value is None or str(value).strip() == '':
            return False
        r = resolve_row(raw_name)
        if not r:
            return False
        nv, note = normalize_value(ws, r, value, rules)
        if note:
            notes.append((r, item_name(ws, r), note))
        if nv is None:
            return True   # 「子行に無」等: 書かないのが正解なので成功扱い
        ws.cell(r, COL_INPUT).value = nv
        return True

    audit = {'name_id': client_id, 'basic': 0, 'measurements_matched': [], 'measurements_unmatched': [],
             'questionnaire_matched': 0, 'questionnaire_unmatched': []}

    # 1) 名前欄(G1)＝識別番号(仮名ID)【確定ルール】・No.0 識別番号
    ws.cell(1, COL_INPUT).value = client_id
    put('識別番号', client_id)

    # 2) 基本情報
    prof = data.get('profile') or {}
    for k, laif in BASIC.items():
        if k in prof and put(laif, prof[k]):
            audit['basic'] += 1

    # 3) 健診 measurements(名前正規化一致＋別名)
    for m in (data.get('measurements') or []):
        if not norm_exact(m.get('name')):
            continue
        if put(m.get('name'), m.get('value')):
            audit['measurements_matched'].append(m.get('name'))
        else:
            audit['measurements_unmatched'].append(m.get('name'))

    # 4) 問診(LAiF項目名で正規化一致)
    for k, v in (data.get('questionnaire') or {}).items():
        if put(k, v):
            audit['questionnaire_matched'] += 1
        else:
            audit['questionnaire_unmatched'].append(k)

    # 5) 親子ルール → 6) 全数検証 → 違反ゼロのときだけ保存
    _apply_parent_child(ws, rules, notes)
    violations = validate(ws, rules)
    if not _report(ws, notes, violations, audit, a.out, warns):
        return 1
    wb.save(a.out)
    print('※未一致項目は空のまま(捏造ゼロ)。写像は §4.1.1 の対応表を拡充して増やす。', file=sys.stderr)
    return 0


def main():
    ap = argparse.ArgumentParser(description='LAiF 上り入力フォームの生成/修復/検証')
    ap.add_argument('--template', help='LAiF 正式フォーム(空)')
    ap.add_argument('--input', help='data.json')
    ap.add_argument('--out', help='出力 xlsx')
    ap.add_argument('--repair', metavar='XLSX', help='既存ファイルを同一ルールで修復して --out へ')
    ap.add_argument('--check', metavar='XLSX', help='既存ファイルを入力規則で全数検証(書き換えない)')
    ap.add_argument('--client-id', help='--repair 時に識別番号(名前欄/No.0)を差し替える')
    a = ap.parse_args()

    if a.check:
        sys.exit(cmd_check(a.check))
    if a.repair:
        if not a.out:
            ap.error('--repair には --out が必要です')
        sys.exit(cmd_repair(a.repair, a.out, a.client_id))
    if not (a.template and a.input and a.out):
        ap.error('--template --input --out（または --repair/--check）を指定してください')
    sys.exit(cmd_build(a))


if __name__ == '__main__':
    main()
